from datetime import datetime, date, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import func, desc
from sqlalchemy.orm import Session
import calendar
import io

from database import get_db
from auth import get_current_user, require_min_role
from models.usuario import Usuario, RolEnum
from models.expediente import Expediente, TipoTraficoCampoEnum, CanalEnum
from models.sesion_trabajo import SesionTrabajo, EstadoSesionEnum
from models.objetivo_mes import ObjetivoMes
from models.tipo_dua import TipoDua
from models.evaluaciones_bonus import EvaluacionBonus
from services.calculo_kpis import calcular_kpis_operario, dias_laborables_mes

router = APIRouter(prefix="/dashboard", tags=["dashboard"])


# ── helpers ──────────────────────────────────────────────────────────────────

def _operarios_filtrados(db: Session, sede: Optional[str] = None):
    q = db.query(Usuario).filter(Usuario.activo == True, Usuario.rol == RolEnum.operario)
    if sede and sede != "todas":
        q = q.filter(Usuario.sede == sede)
    return q.all()


def _exps_periodo(db: Session, año: int, mes: int, ids_operario: list):
    inicio = datetime(año, mes, 1)
    fin = datetime(año + 1, 1, 1) if mes == 12 else datetime(año, mes + 1, 1)
    return (
        db.query(Expediente)
        .filter(
            Expediente.operario_id.in_(ids_operario),
            Expediente.fecha_apertura_dossier >= inicio,
            Expediente.fecha_apertura_dossier < fin,
        )
        .all()
    )


def _dias_hab_transcurridos(año: int, mes: int) -> int:
    hoy = date.today()
    if hoy.year != año or hoy.month != mes:
        return dias_laborables_mes(año, mes)
    return sum(
        1 for d in range(1, hoy.day + 1)
        if date(año, mes, d).weekday() < 5
    )


def _fmt_tiempo(minutos: float) -> str:
    if minutos is None:
        return None
    if minutos < 60:
        return f"{int(minutos)}min"
    h = int(minutos // 60)
    m = int(minutos % 60)
    return f"{h}h {m}min" if m else f"{h}h"


# ── kpis-globales ─────────────────────────────────────────────────────────────

@router.get("/kpis-globales")
def kpis_globales(
    año: int = Query(default=None),
    mes: int = Query(default=None),
    sede: Optional[str] = Query(default=None),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_min_role(RolEnum.coordinador)),
):
    now = datetime.now()
    año = año or now.year
    mes = mes or now.month

    operarios = _operarios_filtrados(db, sede)
    ids = [op.id for op in operarios]

    exps = _exps_periodo(db, año, mes, ids)

    # UPs y objetivo
    ups_totales = sum(e.up_calculadas or 0 for e in exps)
    objetivos = db.query(ObjetivoMes).filter(
        ObjetivoMes.operario_id.in_(ids),
        ObjetivoMes.año == año,
        ObjetivoMes.mes == mes,
    ).all()
    ups_objetivo = sum(o.objetivo_up for o in objetivos)
    pct_rendimiento = round(ups_totales / ups_objetivo * 100, 1) if ups_objetivo > 0 else 0.0

    # Mes anterior
    mes_ant = mes - 1 if mes > 1 else 12
    año_ant = año if mes > 1 else año - 1
    exps_ant = _exps_periodo(db, año_ant, mes_ant, ids)
    ups_ant = sum(e.up_calculadas or 0 for e in exps_ant)
    objs_ant = db.query(ObjetivoMes).filter(
        ObjetivoMes.operario_id.in_(ids),
        ObjetivoMes.año == año_ant,
        ObjetivoMes.mes == mes_ant,
    ).all()
    ups_obj_ant = sum(o.objetivo_up for o in objs_ant)
    pct_rendimiento_ant = round(ups_ant / ups_obj_ant * 100, 1) if ups_obj_ant > 0 else 0.0

    # Proyección fin de mes
    dias_hab_total = dias_laborables_mes(año, mes)
    dias_hab_trans = _dias_hab_transcurridos(año, mes)
    ups_proyectadas = None
    if dias_hab_trans > 4:
        ups_proyectadas = round(ups_totales / dias_hab_trans * dias_hab_total, 1)

    # Cronómetros activos hoy
    hoy_inicio = datetime(now.year, now.month, now.day)
    sesiones_activas = db.query(SesionTrabajo).filter(
        SesionTrabajo.operario_id.in_(ids),
        SesionTrabajo.estado == EstadoSesionEnum.activa,
        SesionTrabajo.inicio >= hoy_inicio,
    ).all()
    operarios_con_cronometro = len(set(s.operario_id for s in sesiones_activas))

    # Sin actividad hoy
    ids_activos_hoy = set(
        e.operario_id for e in exps
        if e.fecha_apertura_dossier and e.fecha_apertura_dossier.date() == now.date()
    )
    operarios_sin_actividad_hoy = len([op for op in operarios if op.id not in ids_activos_hoy])

    # Desglose expedientes
    exp_abiertos = sum(1 for e in exps if e.fecha_apertura_dossier and not e.fecha_levante)
    exp_cerrados = sum(1 for e in exps if e.fecha_levante)
    exp_en_facturacion = sum(1 for e in exps if e.fecha_levante and not e.fecha_envio_facturacion)

    # Factor K por operario → distribución
    verde = naranja = rojo = 0
    for op in operarios:
        op_exps = [e for e in exps if e.operario_id == op.id]
        op_ups = sum(e.up_calculadas or 0 for e in op_exps)
        op_obj_row = next((o for o in objetivos if o.operario_id == op.id), None)
        op_obj = op_obj_row.objetivo_up if op_obj_row else 0
        if op_obj > 0:
            k = op_ups / op_obj
            if k >= 1.0:
                verde += 1
            elif k >= 0.85:
                naranja += 1
            else:
                rojo += 1

    # Tiempos medios
    tiempos_respuesta = [
        (e.fecha_apertura_dossier - e.fecha_recepcion_correo).total_seconds() / 60
        for e in exps
        if e.fecha_recepcion_correo and e.fecha_apertura_dossier
    ]
    tmr_min = round(sum(tiempos_respuesta) / len(tiempos_respuesta), 1) if tiempos_respuesta else None

    tiempos_tmt = [
        (e.fecha_envio_aduana - e.fecha_apertura_dossier).total_seconds() / 60
        for e in exps
        if e.fecha_apertura_dossier and e.fecha_envio_aduana
    ]
    tmt_min = round(sum(tiempos_tmt) / len(tiempos_tmt), 1) if tiempos_tmt else None

    return {
        "año": año, "mes": mes,
        "ups_totales": round(ups_totales, 2),
        "ups_objetivo": round(ups_objetivo, 2),
        "ups_proyectadas": ups_proyectadas,
        "pct_rendimiento": pct_rendimiento,
        "pct_rendimiento_mes_anterior": pct_rendimiento_ant,
        "operarios_activos": len(operarios),
        "operarios_con_cronometro": operarios_con_cronometro,
        "operarios_sin_actividad_hoy": operarios_sin_actividad_hoy,
        "expedientes_total": len(exps),
        "expedientes_abiertos": exp_abiertos,
        "expedientes_cerrados": exp_cerrados,
        "expedientes_en_facturacion": exp_en_facturacion,
        "distribucion_factor_k": {"verde": verde, "naranja": naranja, "rojo": rojo},
        "tiempo_respuesta_medio_min": tmr_min,
        "tiempo_respuesta_medio_fmt": _fmt_tiempo(tmr_min),
        "tiempo_tramitacion_medio_min": tmt_min,
        "tiempo_tramitacion_medio_fmt": _fmt_tiempo(tmt_min),
        "dias_hab_transcurridos": dias_hab_trans,
        "dias_hab_totales": dias_hab_total,
    }


# ── alertas ───────────────────────────────────────────────────────────────────

@router.get("/alertas")
def alertas_dashboard(
    sede: Optional[str] = Query(default=None),
    año: int = Query(default=None),
    mes: int = Query(default=None),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_min_role(RolEnum.coordinador)),
):
    now = datetime.now()
    año = año or now.year
    mes = mes or now.month
    hoy = date.today()
    alertas = []
    operarios = _operarios_filtrados(db, sede)
    ids = [op.id for op in operarios]

    # Expedientes sin asignar
    sin_asignar = db.query(Expediente).filter(Expediente.operario_id == None).count()
    if sin_asignar > 0:
        alertas.append({
            "tipo": "sin_asignar", "severidad": "alta",
            "titulo": "Expedientes sin asignar",
            "descripcion": f"{sin_asignar} expediente{'s' if sin_asignar > 1 else ''} pendiente{'s' if sin_asignar > 1 else ''} de asignación.",
            "link": "/expedientes",
        })

    # Sin actividad 2 días hábiles
    dias_atras, dias_hab = 0, 0
    while dias_hab < 2:
        dias_atras += 1
        d = hoy - timedelta(days=dias_atras)
        if d.weekday() < 5:
            dias_hab += 1
    fecha_umbral = datetime(d.year, d.month, d.day)
    for op in operarios:
        count = db.query(Expediente).filter(
            Expediente.operario_id == op.id,
            Expediente.created_at >= fecha_umbral,
        ).count()
        if count == 0:
            alertas.append({
                "tipo": "sin_actividad", "severidad": "media",
                "titulo": "Operario sin actividad",
                "descripcion": f"{op.nombre} {op.apellidos} no ha registrado actividad en los últimos 2 días hábiles.",
                "link": f"/expedientes?operario_id={op.id}",
            })

    # Objetivo en riesgo (antes del día 20)
    if hoy.year == año and hoy.month == mes and hoy.day < 20:
        exps = _exps_periodo(db, año, mes, ids)
        ups_acum = sum(e.up_calculadas or 0 for e in exps)
        dias_hab_trans = _dias_hab_transcurridos(año, mes)
        dias_hab_total = dias_laborables_mes(año, mes)
        if dias_hab_trans > 4:
            ups_proy = ups_acum / dias_hab_trans * dias_hab_total
            objs = db.query(ObjetivoMes).filter(
                ObjetivoMes.operario_id.in_(ids),
                ObjetivoMes.año == año,
                ObjetivoMes.mes == mes,
            ).all()
            ups_obj = sum(o.objetivo_up for o in objs)
            if ups_obj > 0 and ups_proy / ups_obj < 0.85:
                pct = round(ups_proy / ups_obj * 100, 0)
                alertas.append({
                    "tipo": "objetivo_riesgo", "severidad": "informativa",
                    "titulo": "Objetivo del mes en riesgo",
                    "descripcion": f"Al ritmo actual el equipo cerrará en {ups_proy:.0f} UPs ({pct:.0f}% del objetivo).",
                    "link": "/equipo",
                })

    # Tiempo de respuesta crítico (últimos 3 días)
    hace3 = datetime(hoy.year, hoy.month, hoy.day) - timedelta(days=3)
    exps_recientes = db.query(Expediente).filter(
        Expediente.operario_id.in_(ids),
        Expediente.fecha_apertura_dossier >= hace3,
        Expediente.fecha_recepcion_correo != None,
    ).all()
    tmrs = [
        (e.fecha_apertura_dossier - e.fecha_recepcion_correo).total_seconds() / 60
        for e in exps_recientes
        if e.fecha_apertura_dossier and e.fecha_recepcion_correo
    ]
    if tmrs:
        tmr_medio = sum(tmrs) / len(tmrs)
        if tmr_medio > 240:
            alertas.append({
                "tipo": "tiempo_respuesta", "severidad": "media",
                "titulo": "Tiempo de respuesta crítico",
                "descripcion": f"El tiempo de respuesta medio está en {_fmt_tiempo(tmr_medio)} (últimos 3 días). Por encima del objetivo.",
                "link": "/equipo",
            })

    # Evaluaciones pendientes (> 3 días sin tocar)
    try:
        hace3_dt = datetime.now() - timedelta(days=3)
        eval_pendientes = db.query(EvaluacionBonus).filter(
            EvaluacionBonus.estado == "evaluacion_dir",
            EvaluacionBonus.fecha_inicio_eval_dir <= hace3_dt,
        ).count()
        if eval_pendientes > 0:
            alertas.append({
                "tipo": "evaluaciones", "severidad": "informativa",
                "titulo": "Evaluaciones pendientes",
                "descripcion": f"{eval_pendientes} evaluación{'es' if eval_pendientes > 1 else ''} lleva{'n' if eval_pendientes > 1 else ''} más de 3 días esperando revisión.",
                "link": "/evaluaciones-bonus",
            })
    except Exception:
        pass

    return alertas


# ── evolución ─────────────────────────────────────────────────────────────────

@router.get("/evolucion")
def evolucion(
    meses: int = Query(default=6),
    sede: Optional[str] = Query(default=None),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_min_role(RolEnum.coordinador)),
):
    now = datetime.now()
    operarios = _operarios_filtrados(db, sede)
    ids = [op.id for op in operarios]

    resultado = []
    for i in range(meses - 1, -1, -1):
        mes = now.month - i
        año = now.year
        while mes <= 0:
            mes += 12
            año -= 1

        exps = _exps_periodo(db, año, mes, ids)
        ups = sum(e.up_calculadas or 0 for e in exps)
        objs = db.query(ObjetivoMes).filter(
            ObjetivoMes.operario_id.in_(ids),
            ObjetivoMes.año == año,
            ObjetivoMes.mes == mes,
        ).all()
        ups_obj = sum(o.objetivo_up for o in objs)

        exp_export = sum(1 for e in exps if e.tipo_trafico == TipoTraficoCampoEnum.exportacion)
        exp_import = sum(1 for e in exps if e.tipo_trafico == TipoTraficoCampoEnum.importacion)
        exp_especial = sum(1 for e in exps if e.tipo_trafico == TipoTraficoCampoEnum.regimen_especial)

        resultado.append({
            "año": año, "mes": mes,
            "ups_producidas": round(ups, 2),
            "ups_objetivo": round(ups_obj, 2),
            "expedientes_total": len(exps),
            "expedientes_exportacion": exp_export,
            "expedientes_importacion": exp_import,
            "expedientes_especial": exp_especial,
        })

    return resultado


# ── distribución ──────────────────────────────────────────────────────────────

@router.get("/distribucion")
def distribucion(
    año: int = Query(default=None),
    mes: int = Query(default=None),
    sede: Optional[str] = Query(default=None),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_min_role(RolEnum.coordinador)),
):
    now = datetime.now()
    año = año or now.year
    mes = mes or now.month

    operarios = _operarios_filtrados(db, sede)
    ids = [op.id for op in operarios]
    exps = _exps_periodo(db, año, mes, ids)
    total = len(exps) or 1

    # Por canal
    canal_counts = {c: sum(1 for e in exps if e.canal_respuesta == c) for c in CanalEnum}
    por_canal = {
        c.value: {"count": canal_counts[c], "pct": round(canal_counts[c] / total * 100, 1)}
        for c in [CanalEnum.verde, CanalEnum.naranja, CanalEnum.rojo]
    }

    # Por tipo DUA
    tipo_counts = {}
    tipo_ups = {}
    for e in exps:
        codigo = e.tipo_dua.codigo if e.tipo_dua else "?"
        tipo_counts[codigo] = tipo_counts.get(codigo, 0) + 1
        tipo_ups[codigo] = tipo_ups.get(codigo, 0.0) + (e.up_calculadas or 0)
    por_tipo = sorted([
        {"codigo": k, "count": v, "pct": round(v / total * 100, 1), "ups": round(tipo_ups.get(k, 0), 2)}
        for k, v in tipo_counts.items()
    ], key=lambda x: x["count"], reverse=True)

    # Por tráfico
    exp_exp = [e for e in exps if e.tipo_trafico == TipoTraficoCampoEnum.exportacion]
    exp_imp = [e for e in exps if e.tipo_trafico == TipoTraficoCampoEnum.importacion]
    ups_exp = sum(e.up_calculadas or 0 for e in exp_exp)
    ups_imp = sum(e.up_calculadas or 0 for e in exp_imp)
    ups_total = ups_exp + ups_imp or 1

    por_trafico = {
        "exportacion": {
            "expedientes": len(exp_exp),
            "ups": round(ups_exp, 2),
            "pct_ups": round(ups_exp / ups_total * 100, 1),
        },
        "importacion": {
            "expedientes": len(exp_imp),
            "ups": round(ups_imp, 2),
            "pct_ups": round(ups_imp / ups_total * 100, 1),
        },
    }

    return {"por_canal": por_canal, "por_tipo": por_tipo, "por_trafico": por_trafico}


# ── expedientes en curso ───────────────────────────────────────────────────────

@router.get("/expedientes-en-curso")
def expedientes_en_curso(
    sede: Optional[str] = Query(default=None),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_min_role(RolEnum.coordinador)),
):
    operarios = _operarios_filtrados(db, sede)
    ids = [op.id for op in operarios]
    op_map = {op.id: f"{op.nombre} {op.apellidos}" for op in operarios}

    exps = (
        db.query(Expediente)
        .filter(
            Expediente.operario_id.in_(ids),
            Expediente.fecha_apertura_dossier != None,
            Expediente.fecha_levante == None,
        )
        .order_by(Expediente.fecha_apertura_dossier)
        .all()
    )

    now = datetime.now()
    result = []
    for e in exps:
        horas = (now - e.fecha_apertura_dossier).total_seconds() / 3600
        if horas < 2:
            alerta = None
            badge = "verde"
        elif horas < 4:
            alerta = None
            badge = "naranja"
        else:
            alerta = True
            badge = "rojo"

        if horas < 1:
            tiempo_fmt = f"{int(horas * 60)}min"
        elif horas < 24:
            h = int(horas)
            m = int((horas - h) * 60)
            tiempo_fmt = f"{h}h {m}min"
        else:
            dias = int(horas / 24)
            tiempo_fmt = f"{dias} días"

        estado = "en tramitación"
        if e.fecha_envio_aduana:
            estado = "esperando levante"
        elif not e.fecha_apertura_dossier:
            estado = "pendiente de datos"

        result.append({
            "id": e.id,
            "numero_expediente": e.numero_expediente,
            "operario": op_map.get(e.operario_id, "—"),
            "operario_id": e.operario_id,
            "tipo_dua": e.tipo_dua.codigo if e.tipo_dua else "—",
            "cliente_nombre": e.cliente_nombre,
            "horas_transcurridas": round(horas, 1),
            "tiempo_fmt": tiempo_fmt,
            "badge": badge,
            "estado": estado,
            "alerta": alerta,
        })

    return result


# ── top clientes ──────────────────────────────────────────────────────────────

@router.get("/top-clientes")
def top_clientes(
    año: int = Query(default=None),
    mes: int = Query(default=None),
    limit: int = Query(default=10),
    sede: Optional[str] = Query(default=None),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_min_role(RolEnum.coordinador)),
):
    now = datetime.now()
    año = año or now.year
    mes = mes or now.month

    operarios = _operarios_filtrados(db, sede)
    ids = [op.id for op in operarios]

    exps = _exps_periodo(db, año, mes, ids)

    agg = {}
    for e in exps:
        c = e.cliente_nombre
        if c not in agg:
            agg[c] = {"expedientes": 0, "ups": 0.0, "verde": 0}
        agg[c]["expedientes"] += 1
        agg[c]["ups"] += e.up_calculadas or 0
        if e.canal_respuesta and e.canal_respuesta.value == "verde":
            agg[c]["verde"] += 1

    # Mes anterior para tendencia
    mes_ant = mes - 1 if mes > 1 else 12
    año_ant = año if mes > 1 else año - 1
    exps_ant = _exps_periodo(db, año_ant, mes_ant, ids)
    agg_ant = {}
    for e in exps_ant:
        c = e.cliente_nombre
        agg_ant[c] = agg_ant.get(c, 0) + (e.up_calculadas or 0)

    sorted_items = sorted(agg.items(), key=lambda x: x[1]["ups"], reverse=True)[:limit]

    result = []
    for i, (nombre, datos) in enumerate(sorted_items, 1):
        ups = round(datos["ups"], 2)
        exp_total = datos["expedientes"]
        ups_ant = agg_ant.get(nombre, 0)
        tendencia = "sube" if ups > ups_ant else ("baja" if ups < ups_ant else "igual")
        pct_verde = round(datos["verde"] / exp_total * 100, 1) if exp_total > 0 else 0
        result.append({
            "posicion": i,
            "cliente": nombre,
            "expedientes": exp_total,
            "ups": ups,
            "pct_canal_verde": pct_verde,
            "tendencia": tendencia,
        })

    return result


# ── proyección ────────────────────────────────────────────────────────────────

@router.get("/proyeccion")
def proyeccion(
    año: int = Query(default=None),
    mes: int = Query(default=None),
    sede: Optional[str] = Query(default=None),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_min_role(RolEnum.coordinador)),
):
    now = datetime.now()
    año = año or now.year
    mes = mes or now.month

    operarios = _operarios_filtrados(db, sede)
    ids = [op.id for op in operarios]
    exps = _exps_periodo(db, año, mes, ids)
    ups_acum = sum(e.up_calculadas or 0 for e in exps)

    objs = db.query(ObjetivoMes).filter(
        ObjetivoMes.operario_id.in_(ids),
        ObjetivoMes.año == año,
        ObjetivoMes.mes == mes,
    ).all()
    ups_obj = sum(o.objetivo_up for o in objs)

    dias_hab_total = dias_laborables_mes(año, mes)
    dias_hab_trans = _dias_hab_transcurridos(año, mes)

    if dias_hab_trans < 5:
        return {
            "disponible": False,
            "mensaje": "Proyección disponible a partir del día 5 del mes.",
            "ups_actuales": round(ups_acum, 2),
            "ups_objetivo": round(ups_obj, 2),
        }

    ups_proy = round(ups_acum / dias_hab_trans * dias_hab_total, 2)
    pct_proy = round(ups_proy / ups_obj * 100, 1) if ups_obj > 0 else 0.0

    return {
        "disponible": True,
        "ups_actuales": round(ups_acum, 2),
        "ups_proyectadas": ups_proy,
        "ups_objetivo": round(ups_obj, 2),
        "porcentaje_proyectado": pct_proy,
        "dias_hab_transcurridos": dias_hab_trans,
        "dias_hab_totales": dias_hab_total,
    }


# ── resumen semanal ───────────────────────────────────────────────────────────

@router.get("/resumen-semanal")
def resumen_semanal(
    año: int = Query(default=None),
    mes: int = Query(default=None),
    sede: Optional[str] = Query(default=None),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_min_role(RolEnum.coordinador)),
):
    now = datetime.now()
    año = año or now.year
    mes = mes or now.month

    operarios = _operarios_filtrados(db, sede)
    ids = [op.id for op in operarios]
    op_map = {op.id: op for op in operarios}
    exps = _exps_periodo(db, año, mes, ids)

    semanas = {}
    for e in exps:
        if not e.fecha_apertura_dossier:
            continue
        iso = e.fecha_apertura_dossier.isocalendar()
        key = iso[1]
        if key not in semanas:
            semanas[key] = {"semana": key, "ups": 0.0, "expedientes": 0,
                             "ups_por_op": {}, "exp_por_op": {}}
        semanas[key]["ups"] += e.up_calculadas or 0
        semanas[key]["expedientes"] += 1
        op_id = e.operario_id
        semanas[key]["ups_por_op"][op_id] = semanas[key]["ups_por_op"].get(op_id, 0) + (e.up_calculadas or 0)
        semanas[key]["exp_por_op"][op_id] = semanas[key]["exp_por_op"].get(op_id, 0) + 1

    result = []
    for key in sorted(semanas.keys()):
        s = semanas[key]
        ups_op = s["ups_por_op"]
        if ups_op:
            mejor_id = max(ups_op, key=lambda x: ups_op[x])
            peor_id = min(ups_op, key=lambda x: ups_op[x])
            mejor = op_map[mejor_id].nombre if mejor_id in op_map else "—"
            peor = op_map[peor_id].nombre if peor_id in op_map else "—"
        else:
            mejor = peor = "—"
        result.append({
            "semana_iso": key,
            "ups": round(s["ups"], 2),
            "expedientes": s["expedientes"],
            "mejor_operario": mejor,
            "peor_operario": peor,
        })

    return result


# ── exportar Excel ────────────────────────────────────────────────────────────

@router.get("/exportar-excel")
def exportar_excel(
    año: int = Query(default=None),
    mes: int = Query(default=None),
    sede: Optional[str] = Query(default=None),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_min_role(RolEnum.coordinador)),
):
    now = datetime.now()
    año = año or now.year
    mes = mes or now.month

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        from fastapi import HTTPException
        raise HTTPException(status_code=500, detail="openpyxl no disponible")

    wb = openpyxl.Workbook()

    HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F3864")
    HEADER_FONT = Font(color="FFFFFF", bold=True)

    def _header(ws, cols):
        ws.append(cols)
        for cell in ws[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT

    # Hoja 1 — KPIs globales
    kpis = kpis_globales(año=año, mes=mes, sede=sede, db=db, current_user=current_user)
    ws1 = wb.active
    ws1.title = "KPIs Globales"
    _header(ws1, ["Métrica", "Valor"])
    rows1 = [
        ("UPs producidas", kpis["ups_totales"]),
        ("UPs objetivo", kpis["ups_objetivo"]),
        ("% Rendimiento", f"{kpis['pct_rendimiento']}%"),
        ("UPs proyectadas fin de mes", kpis["ups_proyectadas"] or "—"),
        ("Operarios activos", kpis["operarios_activos"]),
        ("Cronómetros activos", kpis["operarios_con_cronometro"]),
        ("Sin actividad hoy", kpis["operarios_sin_actividad_hoy"]),
        ("Expedientes total", kpis["expedientes_total"]),
        ("Expedientes abiertos", kpis["expedientes_abiertos"]),
        ("Expedientes cerrados", kpis["expedientes_cerrados"]),
        ("Expedientes en facturación", kpis["expedientes_en_facturacion"]),
        ("Tiempo respuesta medio", kpis["tiempo_respuesta_medio_fmt"] or "Sin datos"),
        ("Tiempo tramitación medio", kpis["tiempo_tramitacion_medio_fmt"] or "Sin datos"),
    ]
    for r in rows1:
        ws1.append(r)

    # Hoja 2 — Ranking operarios
    operarios = _operarios_filtrados(db, sede)
    ws2 = wb.create_sheet("Ranking Operarios")
    _header(ws2, ["Operario", "UPs", "Objetivo", "% Cumpl.", "Factor K", "Expedientes"])
    for op in operarios:
        kpi = calcular_kpis_operario(db, op.id, año, mes)
        ws2.append([
            f"{op.nombre} {op.apellidos}",
            round(kpi.get("up_producidas", 0), 2),
            kpi.get("objetivo_up") or "—",
            f"{kpi.get('pct_cumplimiento') or 0:.1f}%",
            f"{kpi.get('factor_k') or 0:.3f}",
            kpi.get("num_expedientes", 0),
        ])

    # Hoja 3 — Distribución
    dist = distribucion(año=año, mes=mes, sede=sede, db=db, current_user=current_user)
    ws3 = wb.create_sheet("Distribución")
    _header(ws3, ["Canal / Tipo", "Expedientes", "% del total"])
    for canal, d in dist["por_canal"].items():
        ws3.append([f"Canal {canal}", d["count"], f"{d['pct']}%"])
    ws3.append([])
    for t in dist["por_tipo"]:
        ws3.append([t["codigo"], t["count"], f"{t['pct']}%"])

    # Hoja 4 — Top clientes
    clientes = top_clientes(año=año, mes=mes, limit=10, sede=sede, db=db, current_user=current_user)
    ws4 = wb.create_sheet("Top Clientes")
    _header(ws4, ["#", "Cliente", "Expedientes", "UPs", "% Canal Verde", "Tendencia"])
    for c in clientes:
        ws4.append([c["posicion"], c["cliente"], c["expedientes"], c["ups"], f"{c['pct_canal_verde']}%", c["tendencia"]])

    # Hoja 5 — Resumen semanal
    semanas = resumen_semanal(año=año, mes=mes, sede=sede, db=db, current_user=current_user)
    ws5 = wb.create_sheet("Resumen Semanal")
    _header(ws5, ["Semana ISO", "UPs", "Expedientes", "Mejor operario", "Peor operario"])
    for s in semanas:
        ws5.append([s["semana_iso"], s["ups"], s["expedientes"], s["mejor_operario"], s["peor_operario"]])

    # Ajustar anchos
    for ws in [ws1, ws2, ws3, ws4, ws5]:
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    from datetime import datetime as dt
    MESES = ["", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
             "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
    filename = f"GECOTEX_Dashboard_{MESES[mes]}_{año}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
