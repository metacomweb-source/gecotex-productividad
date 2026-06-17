from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime


COLOR_HEADER = "1F3864"
COLOR_VERDE = "27AE60"
COLOR_NARANJA = "E67E22"
COLOR_ROJO = "C0392B"


def _header_style(ws, row, headers):
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(fill_type="solid", fgColor=COLOR_HEADER)
        cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[row].height = 18


def _auto_width(ws):
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)


def generar_informe_productividad(datos_operarios: list, año: int, mes: int) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = f"Productividad {mes:02d}/{año}"

    ws["A1"] = f"GECOTEX — Informe de Productividad {mes:02d}/{año}"
    ws["A1"].font = Font(bold=True, size=14, color=COLOR_HEADER)
    ws.merge_cells("A1:K1")

    headers = ["Operario", "UPs Producidas", "Objetivo UP", "% Cumplimiento", "Factor K",
               "Nº Expedientes", "% Incidencia", "TMT (min)", "TF (horas)", "Tiempo Respuesta (min)", "% Bonus"]
    _header_style(ws, 3, headers)

    for i, op in enumerate(datos_operarios, 4):
        ws.cell(row=i, column=1, value=op.get("operario_nombre", ""))
        ws.cell(row=i, column=2, value=op.get("up_producidas", 0))
        ws.cell(row=i, column=3, value=op.get("objetivo_up", ""))
        pct = op.get("pct_cumplimiento")
        ws.cell(row=i, column=4, value=f"{pct:.1f}%" if pct is not None else "")
        k = op.get("factor_k")
        cell_k = ws.cell(row=i, column=5, value=round(k, 3) if k is not None else "")
        if k is not None:
            if k >= 1.0:
                cell_k.fill = PatternFill(fill_type="solid", fgColor=COLOR_VERDE)
            elif k >= 0.85:
                cell_k.fill = PatternFill(fill_type="solid", fgColor=COLOR_NARANJA)
            else:
                cell_k.fill = PatternFill(fill_type="solid", fgColor=COLOR_ROJO)
        ws.cell(row=i, column=6, value=op.get("num_expedientes", 0))
        ws.cell(row=i, column=7, value=f"{op.get('tasa_incidencia', 0):.1f}%")
        ws.cell(row=i, column=8, value=op.get("tiempo_medio_tramitacion_min", ""))
        ws.cell(row=i, column=9, value=op.get("tiempo_medio_facturacion_horas", ""))
        ws.cell(row=i, column=10, value=op.get("tiempo_medio_respuesta_min", ""))
        ws.cell(row=i, column=11, value=f"{op.get('bonus_individual_pct', 0)*100:.1f}%" if op.get("bonus_individual_pct") is not None else "")

    _auto_width(ws)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def generar_informe_expedientes(expedientes: list) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Expedientes"

    headers = ["Nº Expediente", "Operario", "Cliente", "Tipo DUA", "Tráfico", "Canal",
               "UPs", "Partidas", "F.Recepción", "F.Apertura", "F.Envío Aduana", "F.Levante", "F.Facturación", "Origen"]
    _header_style(ws, 1, headers)

    for i, exp in enumerate(expedientes, 2):
        ws.cell(row=i, column=1, value=exp.get("numero_expediente", ""))
        ws.cell(row=i, column=2, value=exp.get("operario_nombre", ""))
        ws.cell(row=i, column=3, value=exp.get("cliente_nombre", ""))
        ws.cell(row=i, column=4, value=exp.get("tipo_dua_nombre", ""))
        ws.cell(row=i, column=5, value=exp.get("tipo_trafico", ""))
        canal = exp.get("canal_respuesta", "")
        cell_canal = ws.cell(row=i, column=6, value=canal)
        if canal == "verde":
            cell_canal.fill = PatternFill(fill_type="solid", fgColor=COLOR_VERDE)
        elif canal == "naranja":
            cell_canal.fill = PatternFill(fill_type="solid", fgColor=COLOR_NARANJA)
        elif canal == "rojo":
            cell_canal.fill = PatternFill(fill_type="solid", fgColor=COLOR_ROJO)
        ws.cell(row=i, column=7, value=exp.get("up_calculadas", 0))
        ws.cell(row=i, column=8, value=exp.get("num_partidas", 1))
        for col, campo in enumerate([
            "fecha_recepcion_correo", "fecha_apertura_dossier", "fecha_envio_aduana",
            "fecha_levante", "fecha_envio_facturacion"
        ], 9):
            val = exp.get(campo)
            ws.cell(row=i, column=col, value=val.strftime("%d/%m/%Y %H:%M") if val else "")
        ws.cell(row=i, column=14, value=exp.get("origen", ""))

    _auto_width(ws)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


MESES_ES = ["Enero","Febrero","Marzo","Abril","Mayo","Junio","Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]
COLOR_AZUL = "2980B9"
COLOR_GRIS = "BDC3C7"


def generar_informe_bonus_semestral(evaluaciones, factores, config, factor_equipo, año: int, semestre: int) -> BytesIO:
    wb = Workbook()

    # ─── Hoja 1: Resumen Bonus ───────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Resumen Bonus"
    titulo = f"GECOTEX — Bonus Semestre {semestre}/{año}"
    ws1["A1"] = titulo
    ws1["A1"].font = Font(bold=True, size=14, color=COLOR_HEADER)
    ws1.merge_cells("A1:L1")

    if factor_equipo:
        fe_txt = "ACTIVADO (+{}%)".format(round(config.factor_equipo_porcentaje * 100, 1)) if factor_equipo.get("activado") else "No activado"
        ws1["A2"] = f"Factor Equipo: {fe_txt} — {factor_equipo.get('meses_cumplidos', 0)}/{factor_equipo.get('meses_totales', 6)} meses cumplidos"
        ws1["A2"].font = Font(bold=True)

    headers1 = ["Nombre", "Salario", "% Máx. Bonus", "Área 1", "Área 2", "Área 3", "Área 4",
                "Total", "Tramo", "% Bonus", "Factor Equipo", "Bonus Semestral €"]
    _header_style(ws1, 4, headers1)

    total_bonus = 0.0
    for i, ev in enumerate(evaluaciones, 5):
        nombre = f"{ev.empleado.nombre} {ev.empleado.apellidos}" if ev.empleado else "—"
        ws1.cell(row=i, column=1, value=nombre)
        ws1.cell(row=i, column=2, value=ev.salario_bruto_anual or "—")
        ws1.cell(row=i, column=3, value=f"{(ev.pct_maximo_bonus or 0)*100:.0f}%")
        ws1.cell(row=i, column=4, value=round(ev.puntuacion_area1 or 0, 2))
        ws1.cell(row=i, column=5, value=round(ev.puntuacion_area2 or 0, 2))
        ws1.cell(row=i, column=6, value=round(ev.puntuacion_area3 or 0, 2))
        ws1.cell(row=i, column=7, value=round(ev.puntuacion_area4 or 0, 2))
        total = ev.puntuacion_total or 0
        cell_tot = ws1.cell(row=i, column=8, value=round(total, 2))
        if total >= 8.5:
            cell_tot.fill = PatternFill(fill_type="solid", fgColor=COLOR_VERDE)
        elif total >= 7.0:
            cell_tot.fill = PatternFill(fill_type="solid", fgColor=COLOR_NARANJA)
        elif total >= 5.0:
            cell_tot.fill = PatternFill(fill_type="solid", fgColor=COLOR_GRIS)
        ws1.cell(row=i, column=9, value=f"{total:.1f}")
        ws1.cell(row=i, column=10, value=f"{(ev.porcentaje_tramo or 0)*100:.0f}%")
        ws1.cell(row=i, column=11, value="Sí" if ev.factor_equipo_aplicado else "No")
        bonus = ev.bonus_semestral_euros or 0
        ws1.cell(row=i, column=12, value=f"{bonus:,.2f} €".replace(",", "."))
        total_bonus += bonus

    # Fila de totales
    fila_total = len(evaluaciones) + 5
    ws1.cell(row=fila_total, column=1, value="TOTAL EQUIPO").font = Font(bold=True)
    ws1.cell(row=fila_total, column=12, value=f"{total_bonus:,.2f} €".replace(",", ".")).font = Font(bold=True)

    _auto_width(ws1)

    # ─── Hoja 2: Detalle por Empleado ────────────────────────────────────────
    ws2 = wb.create_sheet("Detalle por Empleado")
    factores_por_area = {}
    for f in factores:
        factores_por_area.setdefault(f.area, []).append(f)

    fila = 1
    for ev in evaluaciones:
        nombre = f"{ev.empleado.nombre} {ev.empleado.apellidos}" if ev.empleado else "—"
        ws2.cell(row=fila, column=1, value=nombre).font = Font(bold=True, size=12)
        ws2.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=6)
        fila += 1

        # Área 1
        ws2.cell(row=fila, column=1, value="Área 1 — Productividad DUAs").font = Font(bold=True)
        fila += 1
        for label, val in [
            ("Factor K promedio", ev.factor_k_promedio),
            ("% SLA", ev.pct_sla),
            ("% Registro completo", ev.pct_registro),
            ("Puntuación Área 1", ev.puntuacion_area1),
        ]:
            ws2.cell(row=fila, column=2, value=label)
            ws2.cell(row=fila, column=3, value=round(val, 2) if val is not None else "—")
            fila += 1

        # Áreas 2, 3, 4
        area_nombres = {2: "Calidad Operativa", 3: "Gecotex Corporate", 4: "Digitalización y Adaptación"}
        respuestas_map = {r.factor_id: r for r in (ev.respuestas or [])}
        for area_num in (2, 3, 4):
            ws2.cell(row=fila, column=1, value=f"Área {area_num} — {area_nombres[area_num]}").font = Font(bold=True)
            fila += 1
            _header_style(ws2, fila, ["Factor", "Nota Auto", "Nota Dirección", "Nota Final", "Comentario Auto", "Comentario Dir"])
            fila += 1
            for f in factores_por_area.get(area_num, []):
                r = respuestas_map.get(f.id)
                ws2.cell(row=fila, column=1, value=f.nombre)
                ws2.cell(row=fila, column=2, value=r.nota_auto if r else "—")
                ws2.cell(row=fila, column=3, value=r.nota_dir if r else "—")
                ws2.cell(row=fila, column=4, value=r.nota_final if r else "—")
                ws2.cell(row=fila, column=5, value=r.comentario_auto or "" if r else "")
                ws2.cell(row=fila, column=6, value=r.comentario_dir or "" if r else "")
                fila += 1
            punt = getattr(ev, f"puntuacion_area{area_num}", None)
            ws2.cell(row=fila, column=1, value=f"Puntuación Área {area_num}").font = Font(bold=True)
            ws2.cell(row=fila, column=2, value=round(punt, 2) if punt is not None else "—").font = Font(bold=True)
            fila += 1

        ws2.cell(row=fila, column=1, value="Puntuación Total").font = Font(bold=True)
        ws2.cell(row=fila, column=2, value=round(ev.puntuacion_total, 2) if ev.puntuacion_total else "—").font = Font(bold=True)
        fila += 1
        ws2.cell(row=fila, column=1, value="Bonus Semestral")
        ws2.cell(row=fila, column=2, value=f"{ev.bonus_semestral_euros:,.2f} €".replace(",", ".") if ev.bonus_semestral_euros else "—")
        fila += 2

    _auto_width(ws2)

    # ─── Hoja 3: Configuración Aplicada ──────────────────────────────────────
    ws3 = wb.create_sheet("Configuración Aplicada")
    ws3["A1"] = f"Configuración del período {semestre}/{año}"
    ws3["A1"].font = Font(bold=True, size=13)
    fila = 3
    if config:
        for label, val in [
            ("Antigüedad mínima (meses)", config.antiguedad_minima_meses),
            ("Factor Equipo activo", "Sí" if config.factor_equipo_activo else "No"),
            ("Factor Equipo porcentaje", f"{config.factor_equipo_porcentaje*100:.0f}%"),
            ("Meses mínimos para activar Factor Equipo", config.factor_equipo_meses_minimos),
            ("Peso Área 1", f"{config.peso_area1*100:.0f}%"),
            ("Peso Área 2", f"{config.peso_area2*100:.0f}%"),
            ("Peso Área 3", f"{config.peso_area3*100:.0f}%"),
            ("Peso Área 4", f"{config.peso_area4*100:.0f}%"),
            ("Peso autoevaluación", f"{config.peso_auto_evaluacion*100:.0f}%"),
            ("Peso evaluación dirección", f"{config.peso_dir_evaluacion*100:.0f}%"),
        ]:
            ws3.cell(row=fila, column=1, value=label)
            ws3.cell(row=fila, column=2, value=val)
            fila += 1

    _auto_width(ws3)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def generar_informe_bonus(datos_bonus: list, año: int) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = f"Bonus {año}"

    ws["A1"] = f"GECOTEX — Tabla de Bonus Anual {año}"
    ws["A1"].font = Font(bold=True, size=14, color=COLOR_HEADER)
    ws.merge_cells("A1:I1")

    headers = ["Mes", "Operario", "Antigüedad (m)", "Elegible", "UPs Producidas", "Objetivo UP", "Factor K", "% Bonus Prod.", "% Bonus Indiv."]
    _header_style(ws, 3, headers)

    for i, item in enumerate(datos_bonus, 4):
        mes = item.get("mes", 0)
        ws.cell(row=i, column=1, value=MESES_ES[mes - 1] if 1 <= mes <= 12 else mes)
        ws.cell(row=i, column=2, value=item.get("operario_nombre", ""))
        ws.cell(row=i, column=3, value=item.get("antiguedad_meses", 0))
        ws.cell(row=i, column=4, value="Sí" if item.get("elegible") else "No")
        ws.cell(row=i, column=5, value=round(item.get("up_producidas", 0), 2))
        ws.cell(row=i, column=6, value=round(item.get("objetivo_up", 0), 2))
        k = item.get("factor_k", 0) or 0
        cell_k = ws.cell(row=i, column=7, value=round(k, 3))
        if k >= 1.0:
            cell_k.fill = PatternFill(fill_type="solid", fgColor=COLOR_VERDE)
        elif k >= 0.85:
            cell_k.fill = PatternFill(fill_type="solid", fgColor=COLOR_NARANJA)
        else:
            cell_k.fill = PatternFill(fill_type="solid", fgColor=COLOR_ROJO)
        pct_prod = item.get("porcentaje_bonus_productividad", 0) or 0
        pct_ind = item.get("bonus_individual_pct", 0) or 0
        ws.cell(row=i, column=8, value=f"{pct_prod * 100:.1f}%")
        ws.cell(row=i, column=9, value=f"{pct_ind * 100:.1f}%")

    _auto_width(ws)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
