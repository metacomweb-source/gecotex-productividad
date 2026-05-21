from openpyxl import load_workbook
from typing import Dict, List, Optional, Tuple
from datetime import datetime
from sqlalchemy.orm import Session
from models.expediente import Expediente, CanalEnum, TipoTraficoCampoEnum, OrigenEnum
from models.usuario import Usuario
from models.tipo_dua import TipoDua
from services.calculo_up import calcular_up, calcular_partidas_adicionales

COLUMNAS_SISTEMA = {
    "numero_expediente": ["numero expediente", "expediente", "numero", "nº expediente", "num expediente"],
    "fecha_apertura_dossier": ["fecha apertura", "fecha creacion", "fecha alta", "apertura", "fecha_apertura"],
    "fecha_levante": ["fecha levante", "levante", "fecha despacho", "despacho", "fecha_levante"],
    "operario": ["usuario", "operario", "agente", "tramitador"],
    "tipo_trafico": ["tipo trafico", "trafico", "tipo", "importacion exportacion"],
    "canal_respuesta": ["canal", "canal respuesta", "respuesta"],
    "cliente_nombre": ["cliente", "importador exportador", "razon social", "empresa"],
    "num_partidas": ["partidas", "numero partidas", "nº partidas", "num partidas"],
}


def autodetectar_mapeo(columnas_excel: List[str]) -> Dict[str, Optional[str]]:
    mapeo = {campo: None for campo in COLUMNAS_SISTEMA}
    cols_lower = {c.lower().strip(): c for c in columnas_excel}
    for campo, aliases in COLUMNAS_SISTEMA.items():
        for alias in aliases:
            if alias in cols_lower:
                mapeo[campo] = cols_lower[alias]
                break
    return mapeo


def leer_excel_preview(filepath: str, n_filas: int = 5) -> Tuple[List[str], List[dict], int]:
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], [], 0
    headers = [str(c) if c is not None else f"Col{i}" for i, c in enumerate(rows[0])]
    total = len(rows) - 1
    preview = []
    for row in rows[1:n_filas+1]:
        preview.append({headers[i]: (str(v) if v is not None else "") for i, v in enumerate(row)})
    wb.close()
    return headers, preview, total


def normalizar_canal(valor: str) -> str:
    v = str(valor).lower().strip()
    if "verde" in v or "green" in v or v == "a":
        return "verde"
    if "naranja" in v or "orange" in v or v == "b":
        return "naranja"
    if "rojo" in v or "red" in v or v == "c":
        return "rojo"
    return "pendiente"


def normalizar_trafico(valor: str) -> str:
    v = str(valor).lower().strip()
    if "exp" in v or "export" in v:
        return "exportacion"
    if "imp" in v or "import" in v:
        return "importacion"
    return "exportacion"


def parse_fecha(val) -> Optional[datetime]:
    if val is None or str(val).strip() == "" or str(val) == "None":
        return None
    if isinstance(val, datetime):
        return val
    try:
        return datetime.fromisoformat(str(val))
    except Exception:
        pass
    for fmt in ["%d/%m/%Y %H:%M", "%d/%m/%Y", "%Y-%m-%d"]:
        try:
            return datetime.strptime(str(val).strip(), fmt)
        except Exception:
            pass
    return None


def importar_excel(
    db: Session,
    filepath: str,
    mapeo: Dict[str, Optional[str]],
    accion_duplicados: str = "ignorar",
    usuario_id: int = None,
) -> dict:
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        wb.close()
        return {"importados": 0, "actualizados": 0, "ignorados": 0, "con_error": 0, "errores": []}

    headers = [str(c) if c is not None else f"Col{i}" for i, c in enumerate(rows[0])]
    importados = actualizados = ignorados = 0
    errores = []

    for fila_idx, row in enumerate(rows[1:], start=2):
        row_dict = {headers[i]: (row[i] if i < len(row) else None) for i in range(len(headers))}
        try:
            num_exp_col = mapeo.get("numero_expediente")
            num_exp = str(row_dict.get(num_exp_col, "")).strip() if num_exp_col else ""
            if not num_exp or num_exp == "None":
                errores.append({"fila": fila_idx, "error": "Número de expediente vacío"})
                continue

            existente = db.query(Expediente).filter(Expediente.numero_expediente == num_exp).first()
            if existente:
                if accion_duplicados == "ignorar":
                    ignorados += 1
                    continue

            operario_col = mapeo.get("operario")
            operario_val = str(row_dict.get(operario_col, "")).strip() if operario_col else ""
            operario = None
            if operario_val and operario_val != "None":
                operario = db.query(Usuario).filter(
                    (Usuario.nombre + " " + Usuario.apellidos).ilike(f"%{operario_val}%")
                ).first()
                if not operario:
                    operario = db.query(Usuario).filter(Usuario.email.ilike(f"%{operario_val}%")).first()
            if not operario:
                operario = db.query(Usuario).filter(Usuario.rol == "operario").first()
                if operario_val:
                    errores.append({"fila": fila_idx, "error": f"Operario '{operario_val}' no encontrado, asignado por defecto", "nivel": "advertencia"})

            trafico_col = mapeo.get("tipo_trafico")
            trafico_raw = str(row_dict.get(trafico_col, "exportacion")).strip() if trafico_col else "exportacion"
            tipo_trafico = normalizar_trafico(trafico_raw)

            tipo_dua = db.query(TipoDua).filter(TipoDua.tipo_trafico == tipo_trafico, TipoDua.activo == True).first()

            partidas_col = mapeo.get("num_partidas")
            try:
                num_partidas = int(row_dict.get(partidas_col, 1) or 1) if partidas_col else 1
            except Exception:
                num_partidas = 1

            canal_col = mapeo.get("canal_respuesta")
            canal_raw = str(row_dict.get(canal_col, "pendiente")).strip() if canal_col else "pendiente"
            canal = normalizar_canal(canal_raw)

            cliente_col = mapeo.get("cliente_nombre")
            cliente = str(row_dict.get(cliente_col, "Sin cliente")).strip() if cliente_col else "Sin cliente"
            if not cliente or cliente == "None":
                cliente = "Sin cliente"

            aper_col = mapeo.get("fecha_apertura_dossier")
            lev_col = mapeo.get("fecha_levante")
            fecha_apertura = parse_fecha(row_dict.get(aper_col)) if aper_col else None
            fecha_levante = parse_fecha(row_dict.get(lev_col)) if lev_col else None

            if existente and accion_duplicados == "actualizar":
                existente.canal_respuesta = canal
                if fecha_apertura:
                    existente.fecha_apertura_dossier = fecha_apertura
                if fecha_levante:
                    existente.fecha_levante = fecha_levante
                db.commit()
                actualizados += 1
                continue

            partidas_adic = calcular_partidas_adicionales(tipo_dua, num_partidas) if tipo_dua else 0
            up = calcular_up(tipo_dua, num_partidas, []) if tipo_dua else 1.0

            exp = Expediente(
                numero_expediente=num_exp,
                operario_id=operario.id if operario else 1,
                tipo_dua_id=tipo_dua.id if tipo_dua else 1,
                cliente_nombre=cliente,
                tipo_trafico=tipo_trafico,
                num_partidas=num_partidas,
                canal_respuesta=canal,
                fecha_apertura_dossier=fecha_apertura,
                fecha_levante=fecha_levante,
                servicios_adicionales=[],
                partidas_adicionales_count=partidas_adic,
                up_calculadas=up,
                origen="importacion_excel",
            )
            db.add(exp)
            db.commit()
            importados += 1

        except Exception as e:
            errores.append({"fila": fila_idx, "error": str(e)})
            db.rollback()

    wb.close()
    return {
        "importados": importados,
        "actualizados": actualizados,
        "ignorados": ignorados,
        "con_error": len([e for e in errores if e.get("nivel") != "advertencia"]),
        "errores": errores,
    }
